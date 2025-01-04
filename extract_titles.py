import os
from openpyxl import Workbook, load_workbook

def extract_video_titles(directories, output_excel="video_titles.xlsx"):
    # Video file extensions to filter
    video_extensions = ('.mp4', '.avi', '.mkv', '.mov', '.flv', '.wmv', '.webm')
    
    # Prepare to write to Excel
    if os.path.exists(output_excel):
        # Load existing workbook
        workbook = load_workbook(output_excel)
        sheet = workbook.active
        existing_titles = set()  # Use a set for fast duplicate checks
        for row in sheet.iter_rows(min_row=2, values_only=True):
            existing_titles.add((row[1], row[2]))  # (Video Title, Directory Path)
        next_row = sheet.max_row + 1  # Start writing below the last row
    else:
        # Create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Video Titles"
        # Write headers
        sheet.append(["S.No", "Video Title", "Directory Path"])
        existing_titles = set()
        next_row = 2  # First data row

    # Process each directory
    serial_number = next_row - 1  # Adjust serial number based on current row
    duplicates = []  # Track duplicates for reporting
    for directory in directories:
        if not os.path.isdir(directory):
            print(f"The provided directory '{directory}' does not exist. Skipping.")
            continue

        # Collect video file names
        for file in os.listdir(directory):
            if file.lower().endswith(video_extensions):
                entry = (file, directory)
                if entry in existing_titles:
                    duplicates.append(entry)  # Add to duplicates list
                else:
                    serial_number += 1
                    sheet.append([serial_number, file, directory])
                    existing_titles.add(entry)  # Update the set

    # Save the Excel file
    workbook.save(output_excel)

    # Print duplicates
    if duplicates:
        print("The following duplicates were found and skipped:")
        for title, directory in duplicates:
            print(f"Title: {title}, Directory: {directory}")
    else:
        print("No duplicates were found.")

    print(f"Video titles have been updated in '{output_excel}'.")

# Example usage
directory_paths = input("Enter directory paths separated by commas: ").split(',')
directory_paths = [path.strip() for path in directory_paths]  # Clean up extra spaces
extract_video_titles(directory_paths)
